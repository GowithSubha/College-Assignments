
import openpyxl
import os

# Load the Excel workbook
workbook = openpyxl.load_workbook('E:/Programming Language/College & Tution/VS Code/College/UEM/file_name.xlsx')
sheet = workbook.active

# Get the column containing employee names
file_names = [cell.value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) for cell in row]

# Specify the directory where the files will be created
output_dir = 'E:/Programming Language/College & Tution/VS Code/College/UEM/Java to Python/week 1'

# Create the directory if it doesn't exist
os.makedirs(output_dir, exist_ok=True)

# Initialize counters
created_count = 0
existing_count = 0

# Create a file for each employee
for name in file_names:
    # Customize the file name
    filename = os.path.join(output_dir, f'{name}.ipynb')
    
    # Check if file already exists
    if not os.path.exists(filename):
        with open(filename, 'w') as file:
            file.write(f'')
            # You can include additional information from the Excel sheet if needed
        print(f"File created: {filename}")
        created_count += 1
    else:
        print(f"File already exists: {filename}")
        existing_count += 1

print("File creation process completed.")
print(f"Total files created: {created_count}")
print(f"Total files already existed: {existing_count}")

